from openpyxl import load_workbook
import base64
import logging

logger = logging.getLogger(__name__)


class ExcelParser():
    """Load Excel files.


    Args:
        file_path: Path to the file to load.
    """

    def __init__(self, file_path: str):
        """Initialize with file path."""
        self._file_path = file_path

    def parse(self):

        content = self.excel_to_markdown()
        with open("result.md", "w") as f:
            f.write(content)

        return content

    def excel_to_markdown(self):
        # 加载Excel文件
        wb = load_workbook(self._file_path)
        worksheets = wb.worksheets
        excel_content = []

        for ws_idx, ws in enumerate(worksheets, 1):
            # 添加工作表标题
            excel_content.append("# " + ws.title)
            excel_content.append("")  # 空行分隔

            # 收集所有单元格数据（包含图片位置标记）
            cell_data = {}
            max_row = ws.max_row
            max_col = ws.max_column

            # 初始化单元格数据
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell_data[(row, col)] = {
                        "text": "",
                        "bold": False,
                        "hyperlink": None,
                    }

            # 提取图片并记录位置
            for img in ws._images:
                anchor = img.anchor.to
                row, col = anchor.row, anchor.col  # 获取行列索引
                cell_data[(row, col)][
                    "text"
                ] = f"<img src=\"data:image/{img.format};base64,{base64.b64encode(img.ref.getvalue()).decode('utf-8')}\"/>"

            # 提取单元格文本及样式
            for row in ws.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col
            ):
                for cell in row:
                    row_idx, col_idx = cell.row, cell.column
                    # 如果单元格已有图片，则不覆盖
                    if not cell_data[(row_idx, col_idx)]["text"]:
                        cell_data[(row_idx, col_idx)]["text"] = (
                            str(cell.value) if cell.value is not None else ""
                        )

                    # 记录加粗样式
                    cell_data[(row_idx, col_idx)]["bold"] = (
                        cell.font.bold if cell.font else False
                    )

                    # 记录超链接
                    if cell.hyperlink:
                        cell_data[(row_idx, col_idx)][
                            "hyperlink"
                        ] = cell.hyperlink.target

            # 构建Markdown表格
            # 确定表头行（假设第一行为表头）
            table = []
            # 构建表头分隔线
            header_sep = []
            for col in range(1, max_col + 1):
                header_sep.append("---")
            table_sep = "| " + " | ".join(header_sep) + " |"

            # 填充表格内容
            for row in range(1, max_row + 1):
                row_cells = []
                for col in range(1, max_col + 1):
                    data = cell_data[(row, col)]
                    content = data["text"]

                    # 应用超链接样式
                    if data["hyperlink"]:
                        content = f"[{content}]({data['hyperlink']})"

                    # 应用加粗样式
                    if data["bold"]:
                        content = f"**{content}**"
                    # 换行
                    content = content.replace("\n", "<br />")
                    row_cells.append(content)

                table.append("| " + " | ".join(row_cells) + " |")
                if row == 1:
                    table.append(table_sep)

            # 将表格添加到Markdown内容
            excel_content.extend(table)
            excel_content.append("")
        return "\n".join(excel_content)

# 示例使用
if __name__ == "__main__":
    parser = ExcelParser("../ms_office_old/demo1.xls")
    parser.parse()